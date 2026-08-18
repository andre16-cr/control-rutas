import io
import json
import os
import unicodedata
from datetime import datetime, timezone

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

EXCEL_NAME = "Registro Urgencias.xlsx"
SHEET_NAME = "Ruta"
OUTPUT_FILE = "data.json"

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def clean(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def normalize_text(value):
    text = clean(value).upper()
    text = unicodedata.normalize("NFD", text)
    return "".join(
        ch for ch in text
        if unicodedata.category(ch) != "Mn"
    )


def normalize_state(value):
    state = normalize_text(value)

    if state == "COMPLETADO":
        return "COMPLETADO"

    if state == "TRASLADO":
        return "TRASLADO"

    return "PENDIENTE"


def is_urgent(value):
    return normalize_text(value) in {
        "SI", "YES", "TRUE", "1", "X"
    }


def excel_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()

    text = clean(value)

    for fmt in (
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(
                text, fmt
            ).date().isoformat()
        except ValueError:
            pass

    return text


def get_credentials():
    raw = os.environ["GOOGLE_SERVICE_ACCOUNT"]
    info = json.loads(raw)

    return service_account.Credentials.from_service_account_info(
        info,
        scopes=SCOPES
    )


def find_excel(drive):
    query = (
        f"name = '{EXCEL_NAME}' "
        "and trashed = false"
    )

    result = drive.files().list(
        q=query,
        spaces="drive",
        fields="files(id,name,modifiedTime,mimeType)",
        orderBy="modifiedTime desc",
        pageSize=10,
    ).execute()

    files = result.get("files", [])

    if not files:
        raise RuntimeError(
            f"No se encontró '{EXCEL_NAME}' "
            "compartido con la cuenta de servicio."
        )

    return files[0]


def download_excel(drive, file_id):
    request = drive.files().get_media(
        fileId=file_id
    )

    buffer = io.BytesIO()

    downloader = MediaIoBaseDownload(
        buffer,
        request
    )

    done = False

    while not done:
        _, done = downloader.next_chunk()

    buffer.seek(0)

    return buffer


def read_routes(excel_bytes):
    wb = load_workbook(
        excel_bytes,
        read_only=True,
        data_only=True
    )

    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"No existe la hoja '{SHEET_NAME}'."
        )

    ws = wb[SHEET_NAME]

    headers = [
        clean(c.value)
        for c in ws[1]
    ]

    required = [
        "Fecha",
        "Lugar",
        "Destino",
        "Responsable",
        "Compromiso",
        "Estado",
        "Urgencia",
        "Observaciones",
        "Ubicacion",
    ]

    positions = {
        name: headers.index(name)
        for name in required
        if name in headers
    }

    missing = [
        name
        for name in required
        if name not in positions
    ]

    if missing:
        raise RuntimeError(
            f"Faltan columnas requeridas: {missing}"
        )

    records = []
    seq_by_date = {}

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):
        fecha_raw = row[
            positions["Fecha"]
        ]

        lugar = clean(
            row[positions["Lugar"]]
        )

        destino = clean(
            row[positions["Destino"]]
        )

        responsable = clean(
            row[positions["Responsable"]]
        )

        if (
            fecha_raw in (None, "")
            or not (
                lugar
                or destino
                or responsable
            )
        ):
            continue

        fecha = excel_date(
            fecha_raw
        )

        seq_by_date[fecha] = (
            seq_by_date.get(fecha, 0)
            + 1
        )

        records.append({
            "id": (
                fecha.replace("-", "")
                + "-"
                + str(
                    seq_by_date[fecha]
                ).zfill(3)
            ),
            "fecha": fecha,
            "lugar": lugar,
            "destino": destino,
            "responsable": (
                responsable
                or "SIN ASIGNAR"
            ),
            "compromiso": clean(
                row[
                    positions["Compromiso"]
                ]
            ),
            "estado": normalize_state(
                row[
                    positions["Estado"]
                ]
            ),
            "urgencia": is_urgent(
                row[
                    positions["Urgencia"]
                ]
            ),
            "observaciones": clean(
                row[
                    positions[
                        "Observaciones"
                    ]
                ]
            ),
            "ubicacion": clean(
                row[
                    positions["Ubicacion"]
                ]
            ),
        })

    return records


def main():
    creds = get_credentials()

    drive = build(
        "drive",
        "v3",
        credentials=creds,
        cache_discovery=False
    )

    source = find_excel(drive)

    excel = download_excel(
        drive,
        source["id"]
    )

    records = read_routes(
        excel
    )

    payload = {
        "generated_at":
            datetime.now(
                timezone.utc
            ).isoformat(),

        "source": {
            "name":
                source["name"],

            "modified_time":
                source.get(
                    "modifiedTime",
                    ""
                ),
        },

        "records":
            records,
    }

    with open(
        OUTPUT_FILE,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            payload,
            f,
            ensure_ascii=False,
            indent=2
        )

    print(
        f"OK: {len(records)} "
        f"registros exportados "
        f"a {OUTPUT_FILE}"
    )


if __name__ == "__main__":
    main()
